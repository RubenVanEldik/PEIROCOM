import re

import openpyxl
import pandas as pd
import streamlit as st

import utils
import validate


def _get_hydropower_series(sheet, *, min_row, max_row, min_col, max_col, interval):
    """
    Get a Series of data for a specific range in a sheet
    """
    validate.is_integer(min_row)
    validate.is_integer(min_row)
    validate.is_integer(min_row)
    validate.is_integer(min_row)

    # Create an empty Series
    series = pd.Series(dtype="float64")

    # Get a list with the years
    years = list(sheet.iter_rows(min_row=min_row - 1, max_row=min_row - 1, min_col=min_col, max_col=max_col, values_only=True))[0]

    # Get a nested list with the actual values
    values = list(sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))

    # Loop over all values and add them to the DataFrame
    for day_or_week_number, row in zip(range(1, len(values) + 1), values):
        for year, value in zip(years, row):
            # Create a timestamp of the year and day if the day exists (non leap day years have one row too many)
            if interval == "w":
                timestamp = pd.to_datetime(f"{year}-{day_or_week_number - 1:02}-0", format="%Y-%W-%w", utc=True)
            elif interval == "d":
                timestamp = pd.to_datetime(f"{year}-{day_or_week_number:03}", format="%Y-%j", utc=True)
            else:
                raise ValueError("Interval should be either 'w' or 'd'")

            # Only add the value if the timestamp is in the same year (non leap day years have one row too many)
            if timestamp.year == year:
                series[timestamp] = value

    # Return a sorted Series
    return series.sort_index()


def preprocess_hydropower_data(scenarios):
    """
    Preprocess all hydropower data
    """
    assert validate.is_list_like(scenarios)

    for scenario in scenarios:
        # Get a list of all market nodes with hydropower data
        filepath_hydropower = utils.path("input", "eraa", "Climate Data", f"PEMMDB_XX00_Hydro Inflow_{scenario['year']}")
        filename_regex = r"^PEMMDB_([A-Z]{2}[0-9A-Z]{2})_Hydro Inflow_\d{4}.xlsx$"
        market_nodes = sorted([re.search(filename_regex, filename.name).group(1) for filename in filepath_hydropower.iterdir() if re.search(filename_regex, filename.name)])

        # Loop over each hydropower type
        for hydropower_technology in [{"name": "run_of_river", "sheet_name": "Run-of-River and pondage", "interval": "d"}, {"name": "reservoir", "sheet_name": "Reservoir", "interval": "w"}, {"name": "pumped_storage_open", "sheet_name": "Pump storage - Open Loop", "interval": "w"}, {"name": "pumped_storage_closed", "sheet_name": "Pump Storage - Closed Loop", "interval": "w"}]:
            # Create an empty DataFrame for the capacity
            capacity = pd.DataFrame()

            # Create the directory
            directory = utils.path("input", "scenarios", scenario["name"], "hydropower", hydropower_technology["name"])
            utils.path("input", "scenarios", scenario["name"], "hydropower", hydropower_technology["name"]).mkdir(parents=True, exist_ok=True)

            #
            for market_node in market_nodes:
                with st.spinner(f"Importing {market_node} ({utils.format_str(hydropower_technology['name'])})"):
                    # Load the Excel file and get the sheet for the current hydropower technology
                    wb = openpyxl.load_workbook(filepath_hydropower / f"PEMMDB_{market_node}_Hydro Inflow_{scenario['year']}.xlsx", data_only=True)
                    sheet = wb[hydropower_technology["sheet_name"]]

                    # Retrieve the capacities
                    capacity.loc[market_node, "turbine"] = sheet["C6"].value if sheet["C6"].value is not None else 0
                    capacity.loc[market_node, "pump"] = abs(sheet["C5"].value) if sheet["C5"].value is not None else 0  # Use absolute values as some pump capacities are specified as negative and others as positive values
                    capacity.loc[market_node, "reservoir"] = sheet["C7"].value * 1000 if sheet["C7"].value is not None else 0

                    # Create a DataFrame for the temporal data
                    temporal_data = pd.DataFrame()

                    # Calculate the min and max row
                    min_row = 14
                    max_row = min_row + {"d": 365, "w": 52}[hydropower_technology["interval"]]

                    # Retrieve the temporal data
                    temporal_data["inflow_MWh"] = _get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=16, max_col=51, interval=hydropower_technology["interval"]) * 1000
                    temporal_data["min_generation_MWh"] = _get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=52, max_col=87, interval=hydropower_technology["interval"]) * 1000
                    temporal_data["max_generation_MWh"] = _get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=88, max_col=123, interval=hydropower_technology["interval"]) * 1000
                    temporal_data["min_pumping_MWh"] = _get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=124, max_col=159, interval=hydropower_technology["interval"]) * 1000
                    temporal_data["max_pumping_MWh"] = _get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=160, max_col=195, interval=hydropower_technology["interval"]) * 1000
                    temporal_data["min_generation_MW"] = _get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=196, max_col=231, interval=hydropower_technology["interval"])
                    temporal_data["max_generation_MW"] = _get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=232, max_col=267, interval=hydropower_technology["interval"])
                    temporal_data["min_pumping_MW"] = -_get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=268, max_col=303, interval=hydropower_technology["interval"]).abs()  # Absolute is required as some countries specify negative and other positive values
                    temporal_data["max_pumping_MW"] = -_get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=304, max_col=339, interval=hydropower_technology["interval"]).abs()  # Absolute is required as some countries specify negative and other positive values
                    temporal_data["reservoir_soc"] = _get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=340, max_col=375, interval=hydropower_technology["interval"])
                    temporal_data["min_reservoir_soc"] = _get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=376, max_col=411, interval=hydropower_technology["interval"])
                    temporal_data["max_reservoir_soc"] = _get_hydropower_series(sheet, min_row=min_row, max_row=max_row, min_col=412, max_col=447, interval=hydropower_technology["interval"])

                    # Store the temporal data
                    temporal_data.to_csv(directory / f"{market_node}.csv")

            # Store the capacities
            capacity.to_csv(directory / "capacity.csv")
    st.success("The hydropower data for all market nodes is successfully preprocessed")
